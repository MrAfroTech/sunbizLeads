"""
Create an Excel submission file for Q2.
From repo root (jiraAutomation):
  cd SciCodeBench_Assignment/Q2_interactive
  pip install openpyxl
  python create_submission_excel.py
Output: submission_template.xlsx in this folder (same columns as the CSV).
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

HEADERS = [
    "Prompt",
    "DI",
    "Gemini Response",
    "Json_rubrics",
    "Saved HTML file (URL)",
]

EXAMPLE_ROW = [
    'You are helping me design and implement a small public API for a "user preferences" service. '
    'Requirements (initial): Store and retrieve preferences by user ID and preference key. '
    'Support at least: get(key), set(key, value), delete(key), list_keys(). '
    'Before writing any code: 1) Ask at least one clarifying question. 2) Propose a minimal API and confirm. '
    '3) Provide implementation sketch and one edge case. Proceed step by step.',
    "System Instructions (see PROMPT_AND_SI.md)",
    "[Paste the full Gemini response or conversation summary here]",
    "[Paste or attach the filled cuj_rubric.json with human_rating set for each criterion]",
    "https://drive.google.com/your-saved-html-file-url",
]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q2 Submission"
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    for col, value in enumerate(EXAMPLE_ROW, 1):
        ws.cell(row=2, column=col, value=value)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 45
    out = "submission_template.xlsx"
    wb.save(out)
    print(f"Created {out}")

if __name__ == "__main__":
    main()
